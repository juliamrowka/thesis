from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
# from django.contrib.auth import authenticate, login, logout
# from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import DocumentForm
from .models import Document, MLModel
import openpyxl as op
import os
import pandas as pd
import pickle
# from sklearn import preprocessing
# from sklearn import datasets
from sklearn.preprocessing import StandardScaler, MinMaxScaler, Normalizer
from sklearn.decomposition import PCA
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeRegressor, DecisionTreeClassifier
from sklearn.linear_model import LinearRegression
from sklearn.neighbors import KNeighborsRegressor, KNeighborsClassifier
from sklearn.svm import LinearSVR, LinearSVC
from sklearn.naive_bayes import CategoricalNB
from sklearn.model_selection import cross_val_score, cross_validate
from sklearn.metrics import recall_score, accuracy_score, precision_score, max_error, r2_score, mean_squared_error, d2_absolute_error_score, explained_variance_score, mean_absolute_error, f1_score
import numpy as np

def model_form_upload(request):
    """
    show upload form
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            form = DocumentForm(request.POST, request.FILES)
            filename = "documents/user_" + str(request.user.id) + "/data/" + str(request.FILES['document'])
            files = Document.objects.filter(document = filename, user = request.user)
            if form.is_valid() and files.count() == 0:
                newdoc = Document(document = request.FILES['document'], description = request.POST['description'])
                newdoc.user = request.user
                newdoc.save()
                messages.success(request, "You have successfully uploaded file!")
                request.session['filename'] = filename
                # request.session['my_pipeline'] = []
                return redirect('experiment')
            elif files.count() > 0:
                messages.success(request, "A file with that name already exists.")
        else:
            form = DocumentForm()
            # return render(request, 'upload.html', {'form': form})
        documents = Document.objects.filter(user=request.user)
        return render(request, 'upload.html', {'documents': documents, 'form': form})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def choose_file(request, pk):
    """
    description
    """
    if request.user.is_authenticated:
        choosen_file = Document.objects.get(id=pk)
        filename = choosen_file.document
        messages.success(request, "You have successfully chosen file!")
        request.session['filename'] = str(filename)
        if 'max_column' in request.session: del request.session['max_column']
        if 'my_column_transformer' in request.session: del request.session['my_column_transformer']
        if 'my_pipeline' in request.session: del request.session['my_pipeline']
        if 'my_estimator' in request.session: del request.session['my_estimator']
        if 'evaluation' in request.session: del request.session['evaluation']
        # print(filename)
        return redirect('experiment')
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def delete_file(request, pk):
    """
    description
    """
    if request.user.is_authenticated:
        to_delete = get_object_or_404(Document, pk=pk)
        print(f'File to delete: {to_delete}')
        filename = str(to_delete.document)
        print(f'Path of file to delete: {filename}')
        if to_delete and filename:
            to_delete.delete()
            os.remove(filename)
            print('file removed')
            messages.success(request, "You have successfully removed file!")
        if filename == request.session['filename']:
            del request.session['filename']
            if 'max_column' in request.session: del request.session['max_column']
            if 'my_column_transformer' in request.session: del request.session['my_column_transformer']
            if 'my_pipeline' in request.session: del request.session['my_pipeline']
            if 'my_estimator' in request.session: del request.session['my_estimator']
            if 'evaluation' in request.session: del request.session['evaluation']
            print('session filename')
        return redirect('experiment')
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')

def experiment(request):
    """
    description
    """
    if request.user.is_authenticated:
        if 'filename' in request.session:
            if len(request.session['filename']) > 0:
                excel_name = str(request.session['filename'])
                excel_data = list()
                wb = op.load_workbook(request.session['filename'])
                worksheet = wb.active
                request.session['max_column'] = list(range(1,worksheet.max_column + 1))
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)
                context = {'excel_data': excel_data, 'excel_name': excel_name}
                if 'my_column_transformer' in request.session:
                    my_column_transformer = request.session['my_column_transformer']
                    context.update({'my_column_transformer': my_column_transformer})
                if 'my_pipeline' in request.session:
                    my_pipeline = request.session['my_pipeline']
                    context.update({'my_pipeline': my_pipeline})
                if 'my_estimator' in request.session:
                    my_estimator = request.session['my_estimator'] 
                    context.update({'my_estimator': my_estimator})
                if 'evaluation' in request.session:
                    evaluation = request.session['evaluation'] 
                    context.update({'evaluation': evaluation})
                return render(request, 'experiment.html', context)
            else:
                return render(request, 'experiment.html', {})
        else:
            return render(request, 'experiment.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')

def transformer(request):
    """
    description
    """
    if request.user.is_authenticated:
        return render(request, 'transformer.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')

def std(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            column = request.POST['choosen_column']
            if 'my_column_transformer' not in request.session:
                request.session['my_column_transformer'] = [['StandardScaler', ('Chosen column: ', int(column))]]
                # print('just have created pipeline')
            else:
                pipe = request.session['my_column_transformer']
                pipe.append(['StandardScaler', ('Chosen column: ', int(column))])
                request.session['my_column_transformer'] = pipe

                # print(f'pipe exists: {pipe}')
                # print(f'pipeline in session {request.session["my_pipeline"]}')
            return redirect('transformer')
        elif 'max_column' in request.session:
            max_column = request.session['max_column']
            return render(request, 'preprocessing/std.html', {'max_column': max_column})
        else:
            messages.success(request, "You need to choose data file first")
        return redirect('upload')
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def minmax(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            column = request.POST['choosen_column']
            if 'my_column_transformer' not in request.session:
                request.session['my_column_transformer'] = [['MinMaxScaler', ('Chosen column: ', int(column))]]
                # print('no pipeline')
            else:
                pipe = request.session['my_column_transformer']
                pipe.append(['MinMaxScaler', ('Chosen column: ', int(column))])
                request.session['my_column_transformer'] = pipe

                # print(f'pipe exists: {pipe}')
                # print(f'pipeline in session {request.session["my_pipeline"]}')
            return redirect('transformer')
        elif 'max_column' in request.session:
            max_column = request.session['max_column']
            # print(type(max_column))
            return render(request, 'preprocessing/minmax.html', {'max_column': max_column})
        else:
            messages.success(request, "You need to choose data file first")
        return redirect('upload')
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def norm(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            norm_type = request.POST['norm_type']
            if 'my_pipeline' not in request.session:
                request.session['my_pipeline'] = [['Normalizer', ('Norm type: ', norm_type)]]
                # print('no pipeline')
            else:
                pipe = request.session['my_pipeline']
                pipe.append(['Normalizer', ('Norm type: ', norm_type)])
                request.session['my_pipeline'] = pipe

                # print(f'pipe exists: {pipe}')
                # print(f'pipeline in session {request.session["my_pipeline"]}')
            return redirect('transformer')
        elif 'max_column' in request.session:
            return render(request, 'preprocessing/norm.html', {})
        else:
            messages.success(request, "You need to choose data file first")
        return redirect('upload')
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def pca(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            parameter_n = request.POST['parameter_n']
            if 'my_pipeline' not in request.session:
                request.session['my_pipeline'] = [['PCA', ('Chosen n parameter: ', int(parameter_n))]]
                # print('no pipeline')
            else:
                pipe = request.session['my_pipeline']
                pipe.append(['PCA', ('Chosen n parameter: ', int(parameter_n))])
                request.session['my_pipeline'] = pipe

                # print(f'pipe exists: {pipe}')
                # print(f'pipeline in session {request.session["my_pipeline"]}')
            return redirect('transformer')
        elif 'max_column' in request.session:
            n_component = request.session['max_column']
            return render(request, 'preprocessing/pca.html', {'n_component': n_component})
        else:
            messages.success(request, "You need to choose data file first")
        return redirect('upload')
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def estimator(request):
    """
    description
    """
    if request.user.is_authenticated:
        return render(request, 'estimator.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def ord_least_squares(request):
    """
    description
    """
    # if request.user.is_authenticated:
    #     return render(request, 'ordinary-least-squares.html', {})
    # else:
    #     messages.success(request, "You need to log in first")
    #     return redirect('home')
    if request.user.is_authenticated:
        if request.method == 'POST':
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('LinearRegression', 'reg'), [('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'regression/ordinary_least_squares.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def svm_regression(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            epsilon = request.POST['epsilon']
            C_parameter = request.POST['C_parameter']
            intercept_scaling = request.POST['intercept_scaling']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('LinearSVR', 'reg'), [('Epsilon: ', epsilon), ('Regularization parameter C: ', C_parameter), ('Intercept scaling: ', intercept_scaling), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'regression/svm_regression.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def nn_regression(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            neighbors = request.POST['neighbors']
            # weight = request.POST['weight']
            # p_parameter = request.POST['p_parameter']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('KNeighborsRegressor', 'reg'), [('Number of neighbors: ', neighbors), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'regression/nn_regression.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def dt_regression(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            criterion = request.POST['criterion']
            max_leaf_nodes = request.POST['max_leaf_nodes']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('DecisionTreeRegressor', 'reg'), [('Criterion: ', criterion), ('Max leaf nodes: ', max_leaf_nodes), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'regression/dt_regression.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def categorical_nb(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            alpha = request.POST['alpha']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('CategoricalNB', 'clf'), [('Alpha: ', alpha), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'classification/categorical_nb.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def svm_classification(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            c_parameter = request.POST['c_parameter']
            if request.POST['class_weight'] == 'None': class_weight = None
            else: class_weight = request.POST['class_weight']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('LinearSVC', 'clf'), [('C parameter: ', c_parameter), ('Class weight: ', class_weight), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'classification/svm_classification.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def nn_classification(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            neighbors = request.POST['neighbors']
            # weight = request.POST['weight']
            # p_parameter = request.POST['p_parameter']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('KNeighborsClassifier', 'clf'), [('Number of neighbors: ', neighbors), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'classification/nn_classification.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')

def dt_classification(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            criterion = request.POST['criterion']
            max_leaf_nodes = request.POST['max_leaf_nodes']
            column = request.POST['choosen_column']
            request.session['my_estimator'] = [('DecisionTreeClassifier', 'clf'), [('Criterion: ', criterion), ('Max leaf nodes: ', max_leaf_nodes), ('Column of target values: ', int(column))]]
            return redirect('experiment')
        else:
            return render(request, 'classification/dt_classification.html', {'max_column': request.session['max_column']})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')
    
def evaluation(request):
    """
    description
    """
    if request.user.is_authenticated:
        return render(request, 'evaluation.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')

def random_split(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            test_size = request.POST['test_size']
            request.session['evaluation'] = ['train_test_split', [('test_size: ', test_size)]]
            return redirect('experiment')
        else:
            return render(request, 'evaluation/random_split.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')  

def cross_validation(request):
    """
    description
    """
    if request.user.is_authenticated:
        if request.method == 'POST':
            cv = request.POST['cv']
            request.session['evaluation'] = ['cross_validate', [('cv: ', cv)]]
            return redirect('experiment')
        else:
            return render(request, 'evaluation/cross_validation.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')  

def compute(request):
    """
    description
    """
    if request.user.is_authenticated:
        if 'my_estimator' and 'evaluation' in request.session:
            transformer_list = []
            column_transformer_list = []

            # iteracja po elementach tablicy 'my_column_transformer' w celu odczytania transformacji i przyłączenia jej do listy
            x = 0
            if 'my_column_transformer' in request.session:
                for i in request.session['my_column_transformer']:
                    if i[0] == 'StandardScaler': column_transformer_list.append((f'std_{x}', StandardScaler(), [i[1][1]-1]))
                    if i[0] == 'MinMaxScaler': column_transformer_list.append((f'minmax_{x}', MinMaxScaler(), [i[1][1]-1]))
                    x = x+1

            # iteracja po elementach tablicy 'my_pipeline' w celu odczytania transformacji i przyłączenia jej do listy
            x = 0
            if 'my_pipeline' in request.session:
                for i in request.session['my_pipeline']:
                    if i[0] == 'Normalizer': transformer_list.append((f'norm_{x}', Normalizer(norm=i[1][1])))
                    if i[0] == 'PCA': transformer_list.append((f'pca_{x}', PCA(n_components=i[1][1])))
                    x = x+1
            # print(f'Norma: {transformer_list[0][1].norm}')
            
            est = request.session['my_estimator']

            # odczytywanie wybranego estymatora i przyłączenie go do listy    
            if est[0][0] == 'LinearRegression': transformer_list.append(('linear', LinearRegression()))
            elif est[0][0] == 'LinearSVR': transformer_list.append(('svr', LinearSVR(epsilon=float(est[1][0][1]), C=float(est[1][1][1]), intercept_scaling=float(est[1][2][1]) )))
            elif est[0][0] == 'KNeighborsRegressor': transformer_list.append(('knn_reg', KNeighborsRegressor(n_neighbors=int(est[1][0][1]) )))
            elif est[0][0] == 'DecisionTreeRegressor': transformer_list.append(('tree_reg', DecisionTreeRegressor(criterion=est[1][0][1], max_leaf_nodes=int(est[1][1][1]) )))
            
            elif est[0][0] == 'CategoricalNB': transformer_list.append(('nb', CategoricalNB(alpha=float(est[1][0][1]) )))
            elif est[0][0] == 'LinearSVC': transformer_list.append(('svc', LinearSVC(C=float(est[1][0][1]), class_weight=(est[1][1][1]) )))
            elif est[0][0] == 'KNeighborsClassifier': transformer_list.append(('knn_class', KNeighborsClassifier(n_neighbors=int(est[1][0][1]) )))
            elif est[0][0] == 'DecisionTreeClassifier': transformer_list.append(('tree_class', DecisionTreeClassifier(criterion=est[1][0][1], max_leaf_nodes=int(est[1][1][1]) )))
           
            # print(column_transformer_list)
            # print(transformer_list)

            # stworzenie ColumnTransformer na podstawie listy 'column_transformer_list'
            ct = ColumnTransformer(column_transformer_list, remainder="passthrough")
            # print(type(ct))

            # stworzenie Pipeline z ColumnTransformer oraz dołączenie transformacji i estymatora z listy 'transformer_list'
            pipe = Pipeline([('ct', ct)])
            pipe.steps.extend(transformer_list)
            print(pipe)

            # odczytanie pliku wgranego przez użytkownika oraz wybranej kolumny decyzyjnej - zmiennej celu
            wb = pd.read_excel(io=request.session['filename'], header=0)
            target_column = est[1][len(est[1])-1][1] - 1
            # print(f'target_column: {target_column}')

            # podział danych na x i y (zmienne zależne i niezależna)
            X = wb.iloc[:,[x for x in range(len(wb.columns)) if x!=target_column]]
            y = wb.iloc[:, target_column]
            # print(X.shape, y.shape)

            ev = request.session['evaluation']
            context = {}

            # wybór sposobu podziału danych na zbiór uczący i testujący
            if ev[0] == 'train_test_split':
                test_size = float(ev[1][0][1])
                X_train,X_test,y_train,y_test=train_test_split(X, y, test_size=test_size, random_state=0)
                # print(X_train.shape, y_train.shape)
                # print(X_test.shape, y_test.shape)

                pipe.fit(X_train,y_train)
                y_pred=pipe.predict(X_test)
                # print(y_pred.shape)

                if est[0][1] == 'reg':
                    
                    # print(f'pipe score {pipe.score(X_test,y_test)}')
                    # print(f'd2_absolute_error_score {d2_absolute_error_score(y_test, y_pred)}')
                    # print(f'r2 score {r2_score(y_test, y_pred)}')
                    # print(f'mean_squared_error {mean_squared_error(y_test, y_pred)}')
                    scores1_reg = {
                        'explained variance score': explained_variance_score(y_test, y_pred),
                        'r2 score': r2_score(y_test, y_pred),
                        'neg mean absolute error': mean_absolute_error(y_test, y_pred),
                        'neg mean squared error': mean_squared_error(y_test, y_pred) 
                    }

                    for x in scores1_reg.keys():
                        scores1_reg[x] = round(scores1_reg[x], 4)

                    print(scores1_reg)
                    context = {'scores1_reg': scores1_reg}

                elif est[0][1] == 'clf':
                    # print(np.transpose(np.array([y_test,y_pred])))
                    # print(f'precision score {precision_score(y_test, y_pred, average="macro")}')
                    # print(f'recall score {recall_score(y_test, y_pred, average="macro")}')

                    scores1_clf = {
                        'precision_macro': precision_score(y_test, y_pred, average='macro'),
                        'precision_micro': precision_score(y_test, y_pred, average='micro'),
                        'recall_macro': recall_score(y_test, y_pred, average='macro'),
                        'recall_micro': recall_score(y_test, y_pred, average='micro'),
                        'f1_macro': f1_score(y_test, y_pred, average='macro'),
                        'f1_micro': f1_score(y_test, y_pred, average='micro'),
                        'accuracy': accuracy_score(y_test, y_pred)
                    }

                    for x in scores1_clf.keys():
                        scores1_clf[x] = round(scores1_clf[x], 4)

                    context = {'scores1_clf': scores1_clf}

            elif ev[0] == 'cross_validate':
                cv = int(ev[1][0][1])
                print(cv)

                if est[0][1] == 'reg': 
                    scoring = ['explained_variance', 'r2', 'neg_mean_absolute_error', 'neg_mean_squared_error' ]

                    scores2_reg = cross_validate(pipe, X, y, cv=cv, scoring=scoring)

                    for x in scores2_reg.keys():
                        scores2_reg[x] = [scores2_reg[x].mean(), scores2_reg[x].std()]

                    for x, y in scores2_reg.items():
                        list = []
                        for j in y:
                            list.append(round(j, 4))
                        scores2_reg[x] = list

                    context = {'scores2_reg': scores2_reg}
                    
                    
                    # print(f"Explained variance mean {scores['test_explained_variance'].mean()}")
                    # print(f"Explained variance std {scores['test_explained_variance'].std()}")
                    # print(f"R2 mean {scores['test_r2'].mean()}")
                    # print(f"R2 std {scores['test_r2'].std()}")

                elif est[0][1] == 'clf': 
                    scoring = ['precision_macro', 'precision_micro', 'recall_macro', 'recall_micro', 'f1_macro', 'f1_micro', 'accuracy']

                    scores2_clf = cross_validate(pipe, X, y, cv=cv, scoring=scoring)

                    for x in scores2_clf.keys():
                        # print(x)
                        # print([scores2[x].mean(), scores2[x].std()])
                        scores2_clf[x] = [scores2_clf[x].mean(), scores2_clf[x].std()]
                    
                    # print(scores2_clf)
                    # scores2 = cross_validate(pipe, X, y, cv=cv, scoring=scoring)
                    # scores2['test_precision_macro'] = [scores2['test_precision_macro'].mean(), scores2['test_precision_macro'].std()]
                    # scores2['test_precision_micro'] = [scores2['test_precision_micro'].mean(), scores2['test_precision_micro'].std()]
                    # scores2['test_recall_macro'] = [scores2['test_recall_macro'].mean(), scores2['test_recall_macro'].std()]
                    # scores2['test_recall_micro'] = [scores2['test_recall_micro'].mean(), scores2['test_recall_micro'].std()]
                    # scores2['test_f1_macro'] = [scores2['test_f1_macro'].mean(), scores2['test_f1_macro'].std()]
                    # scores2['test_f1_micro'] = [scores2['test_f1_micro'].mean(), scores2['test_f1_micro'].std()]
                    # scores2['test_accuracy'] = [scores2['test_accuracy'].mean(), scores2['test_accuracy'].std()]

                    for x, y in scores2_clf.items():
                        list = []
                        for j in y:
                            print(j)
                            list.append(round(j, 4))
                        scores2_clf[x] = list

                    context = {'scores2_clf': scores2_clf}
                    print(scores2_clf)
                    
                        
                    
                    # print(f"Precision mean {scores['test_precision_macro'].mean()}")
                    # print(f"Precision std {scores['test_precision_macro'].std()}")
                    # print(f"Recall mean {scores['test_recall_macro'].mean()}")
                    # print(f"Recall std {scores['test_recall_macro'].std()}")
                    # print(f"Accuracy mean {scores['test_accuracy'].mean()}")
                    # print(f"Accuracy std {scores['test_accuracy'].std()}")
                # scores2 = cross_validate(pipe, X, y, cv=cv, scoring=scoring)

                # print(scores2.keys())
                # print(scores2)
            
            if request.method == 'POST':
                filename = request.POST['filename']             
                file = "documents/user_" + str(request.user.id) + "/models/" + filename + ".pickle"
                os.makedirs("documents/user_" + str(request.user.id) + "/models/", exist_ok=True)

                with open(file, 'wb') as f:
                    pickle.dump(pipe, f)

                newmodel = MLModel(file = file)
                newmodel.user = request.user
                newmodel.save()

            return render(request, 'compute.html', context)
        else:
            messages.success(request, "No data")
            return render(request, 'compute.html', {}) 
        # return render(request, 'compute.html', {})
    else:
        messages.success(request, "You need to log in first")
        return redirect('home')