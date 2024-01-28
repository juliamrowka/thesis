from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from .forms import DocumentForm
from .models import Document, MLModel
import openpyxl as op
import os
import pandas as pd
import pickle
from sklearn.preprocessing import StandardScaler, MinMaxScaler, Normalizer
from sklearn.decomposition import PCA
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeRegressor, DecisionTreeClassifier
from sklearn.linear_model import LinearRegression
from sklearn.neighbors import KNeighborsRegressor, KNeighborsClassifier
from sklearn.svm import LinearSVR, LinearSVC
from sklearn.naive_bayes import CategoricalNB
from sklearn.model_selection import cross_validate
from sklearn.metrics import recall_score, accuracy_score, precision_score, r2_score, mean_squared_error, explained_variance_score, mean_absolute_error, f1_score


# functions
def switch_order(arr, a, b):
    tmp = arr[a]
    arr[a] = arr[b]
    arr[b] = tmp

def switch_up(arr, a):
    if a > 0:
        switch_order(arr, a, a-1)

def switch_down(arr, a):
    if a < len(arr)-1:
        switch_order(arr, a, a+1)

# views
def model_form_upload(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            form = DocumentForm(request.POST, request.FILES)
            filename = "documents/user_" + str(request.user.id) + "/data/" + str(request.FILES['document'])
            files = Document.objects.filter(document = filename, user = request.user)
            if form.is_valid() and files.count() == 0:
                newdoc = Document(document = request.FILES['document'], description = request.POST['description'])
                newdoc.user = request.user
                newdoc.save()
                messages.success(request, "Pomyślnie przesłano plik!")
                request.session['filename'] = filename
                return redirect('experiment')
            elif files.count() > 0:
                messages.success(request, "Plik o tej nazwie już istnieje")
        else:
            form = DocumentForm()
        documents = Document.objects.filter(user=request.user)
        return render(request, 'upload.html', {'documents': documents, 'form': form})
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def choose_file(request, pk):
    if request.user.is_authenticated:
        choosen_file = Document.objects.get(id=pk)
        filename = choosen_file.document
        messages.success(request, "Pomyślnie wybrano plik!")
        request.session['filename'] = str(filename)
        if 'max_column' in request.session: del request.session['max_column']
        if 'my_column_transformer' in request.session: del request.session['my_column_transformer']
        if 'my_pipeline' in request.session: del request.session['my_pipeline']
        if 'my_estimator' in request.session: del request.session['my_estimator']
        if 'my_evaluation' in request.session: del request.session['my_evaluation']
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def delete_file(request, pk):
    if request.user.is_authenticated:
        to_delete = get_object_or_404(Document, pk=pk)
        filename = str(to_delete.document)
        if to_delete and filename:
            to_delete.delete()
            os.remove(filename)
            messages.success(request, "Pomyślnie usunięto plik!")
        if filename == request.session['filename']:
            del request.session['filename']
            if 'max_column' in request.session: del request.session['max_column']
            if 'my_column_transformer' in request.session: del request.session['my_column_transformer']
            if 'my_pipeline' in request.session: del request.session['my_pipeline']
            if 'my_estimator' in request.session: del request.session['my_estimator']
            if 'my_evaluation' in request.session: del request.session['my_evaluation']
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def experiment(request):
    if request.user.is_authenticated:
        if 'filename' in request.session:
            if len(request.session['filename']) > 0:
                excel_name = os.path.basename(str(request.session['filename']))
                excel_data = list()
                wb = op.load_workbook(request.session['filename'])
                worksheet = wb.active
                # tworzenie listy indeksów kolumn (cech) - bez ostatniej kolumny (decyzyjnej)
                request.session['max_column'] = list(range(1,worksheet.max_column))
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)
                context = {'excel_data': excel_data, 'excel_name': excel_name}
                if 'my_column_transformer' in request.session:
                    my_column_transformer = request.session['my_column_transformer']
                    context.update({'my_column_transformer': my_column_transformer})
                if 'my_pipeline' in request.session:
                    my_pipeline = request.session['my_pipeline']
                    context.update({'my_pipeline': my_pipeline})
                if 'my_estimator' in request.session:
                    my_estimator = request.session['my_estimator'] 
                    context.update({'my_estimator': my_estimator})
                if 'my_evaluation' in request.session:
                    my_evaluation = request.session['my_evaluation'] 
                    context.update({'my_evaluation': my_evaluation})
                return render(request, 'experiment.html', context)
            else:
                return render(request, 'experiment.html', {})
        else:
            return render(request, 'experiment.html', {})
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def transformer(request):
    if request.user.is_authenticated:
        return render(request, 'transformer.html', {})
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def std(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            column = request.POST['choosen_column']
            if 'my_pipeline' not in request.session:
                request.session['my_pipeline'] = [['StandardScaler', ('Wybrana kolumna: ', int(column))]]
            else:
                pipe = request.session['my_pipeline']
                pipe.append(['StandardScaler', ('Wybrana kolumna: ', int(column))])
                request.session['my_pipeline'] = pipe

            return redirect('transformer')
        elif 'max_column' in request.session:
            max_column = request.session['max_column']
            return render(request, 'preprocessing/std.html', {'max_column': max_column})
        else:
            messages.success(request, "Najpierw wybierz plik")
        return redirect('upload')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def minmax(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            column = request.POST['choosen_column']
            if 'my_pipeline' not in request.session:
                request.session['my_pipeline'] = [['MinMaxScaler', ('Wybrana kolumna: ', int(column))]]
            else:
                pipe = request.session['my_pipeline']
                pipe.append(['MinMaxScaler', ('Wybrana kolumna: ', int(column))])
                request.session['my_pipeline'] = pipe

            return redirect('transformer')
        elif 'max_column' in request.session:
            max_column = request.session['max_column']
            return render(request, 'preprocessing/minmax.html', {'max_column': max_column})
        else:
            messages.success(request, "Najpierw wybierz plik")
        return redirect('upload')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def norm(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            norm_type = request.POST['norm_type']
            if 'my_pipeline' not in request.session:
                request.session['my_pipeline'] = [['Normalizer', ('Norma: ', norm_type)]]
            else:
                pipe = request.session['my_pipeline']
                pipe.append(['Normalizer', ('Norma: ', norm_type)])
                request.session['my_pipeline'] = pipe

            return redirect('transformer')
        elif 'max_column' in request.session:
            max_column = request.session['max_column']
            return render(request, 'preprocessing/norm.html', {'max_column': max_column})
        else:
            messages.success(request, "Najpierw wybierz plik")
        return redirect('upload')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def pca(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            if 'my_pipeline' not in request.session:
                request.session['my_pipeline'] = [['PCA']]
            else:
                pipe = request.session['my_pipeline']
                pipe.append(['PCA'])
                request.session['my_pipeline'] = pipe

            return redirect('transformer')
        elif 'filename' in request.session:
            return render(request, 'preprocessing/pca.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
        return redirect('upload')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def estimator(request):
    if request.user.is_authenticated:
        return render(request, 'estimator.html', {})
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def ord_least_squares(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            request.session['my_estimator'] = [('LinearRegression', 'reg')]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'regression/ordinary_least_squares.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def svm_regression(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            epsilon = request.POST['epsilon']
            C_parameter = request.POST['C_parameter']
            intercept_scaling = request.POST['intercept_scaling']
            request.session['my_estimator'] = [('LinearSVR', 'reg'), [('Parametr epsilon: ', epsilon), ('Parametr regularyzacji C: ', C_parameter), ('Parametr intercept_scaling: ', intercept_scaling)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'regression/svm_regression.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def nn_regression(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            neighbors = request.POST['neighbors']
            request.session['my_estimator'] = [('KNeighborsRegressor', 'reg'), [('Liczba sąsiadów: ', neighbors)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'regression/nn_regression.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')            
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def dt_regression(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            criterion = request.POST['criterion']
            max_leaf_nodes = request.POST['max_leaf_nodes']
            request.session['my_estimator'] = [('DecisionTreeRegressor', 'reg'), [('Kryterium: ', criterion), ('Parametr max_leaf_nodes: ', max_leaf_nodes)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'regression/dt_regression.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')            
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def categorical_nb(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            alpha = request.POST['alpha']
            request.session['my_estimator'] = [('CategoricalNB', 'clf'), [('Parametr alpha: ', alpha)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'classification/categorical_nb.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')    
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def svm_classification(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            c_parameter = request.POST['c_parameter']
            if request.POST['class_weight'] == 'None': class_weight = None
            else: class_weight = request.POST['class_weight']
            request.session['my_estimator'] = [('LinearSVC', 'clf'), [('Parametr regularyzacji C: ', c_parameter), ('Parametr class_weight: ', class_weight)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'classification/svm_classification.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')               
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def nn_classification(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            neighbors = request.POST['neighbors']
            request.session['my_estimator'] = [('KNeighborsClassifier', 'clf'), [('Liczba sąsiadów: ', neighbors)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'classification/nn_classification.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload') 
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def dt_classification(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            criterion = request.POST['criterion']
            max_leaf_nodes = request.POST['max_leaf_nodes']
            request.session['my_estimator'] = [('DecisionTreeClassifier', 'clf'), [('Kryterium: ', criterion), ('Parametr max_leaf_nodes: ', max_leaf_nodes)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'classification/dt_classification.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload') 
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def evaluation(request):
    if request.user.is_authenticated:
        return render(request, 'evaluation.html', {})
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def random_split(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            test_size = request.POST['test_size']
            request.session['my_evaluation'] = ['Train test split', [('Udział zbioru testowego: ', test_size)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'evaluation/random_split.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')             
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')  

def cross_validation(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            cv = request.POST['cv']
            request.session['my_evaluation'] = ['Cross validate', [('Liczba podzbiorów: ', cv)]]
            return redirect('experiment')
        elif 'filename' in request.session:
            return render(request, 'evaluation/cross_validation.html', {})
        else:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')            
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')  

def compute(request):
    if request.user.is_authenticated:
        if 'filename' and 'my_estimator' and 'my_evaluation' in request.session:

            est = request.session['my_estimator']
            ev = request.session['my_evaluation']
            max_c = request.session['max_column']
            context = {}
            transformer_list = []

            # iteracja po elementach tablicy 'my_pipeline' w celu odczytania transformacji i przyłączenia jej do listy
            x = 0
            if 'my_pipeline' in request.session:
                for i in request.session['my_pipeline']:
                    if i[0] == 'StandardScaler':
                        column = i[1][1]-1

                        # brak zmiany kolejności kolumn
                        before = (f'pass_before_{x}', 'passthrough', [v for v in range(column)])
                        trans = (f'std_{x}', StandardScaler(), [column])
                        after = (f'pass_after_{x}', 'passthrough', [v for v in range(column+1, len(max_c))])

                        transformer_list.append((f'ct_std_{x}', ColumnTransformer([before, trans, after])))
              
                    if i[0] == 'MinMaxScaler':
                        column = i[1][1]-1

                        # brak zmiany kolejności kolumn
                        before = (f'pass_before_{x}', 'passthrough', [v for v in range(column)])
                        trans = (f'std_{x}', MinMaxScaler(), [column])
                        after = (f'pass_after_{x}', 'passthrough', [v for v in range(column+1, len(max_c))])

                        transformer_list.append((f'ct_minmax_{x}', ColumnTransformer( [before, trans, after] )))                  

                    if i[0] == 'Normalizer': 
                        transformer_list.append((f'ct_norm_{x}', ColumnTransformer([(f'norm_{x}', Normalizer(norm=i[1][1]), [x for x in range(len(max_c))])])))

                    if i[0] == 'PCA': 
                        transformer_list.append((f'pca_{x}', PCA()))

                    x = x+1
            
            # odczytywanie wybranego estymatora i przyłączenie go do listy    
            if est[0][0] == 'LinearRegression': transformer_list.append(('linear', LinearRegression()))
            elif est[0][0] == 'LinearSVR': transformer_list.append(('svr', LinearSVR(epsilon=float(est[1][0][1]), C=float(est[1][1][1]), intercept_scaling=float(est[1][2][1]) )))
            elif est[0][0] == 'KNeighborsRegressor': transformer_list.append(('knn_reg', KNeighborsRegressor(n_neighbors=int(est[1][0][1]) )))
            elif est[0][0] == 'DecisionTreeRegressor': transformer_list.append(('tree_reg', DecisionTreeRegressor(criterion=est[1][0][1], max_leaf_nodes=int(est[1][1][1]) )))
            
            elif est[0][0] == 'CategoricalNB': transformer_list.append(('nb', CategoricalNB(alpha=float(est[1][0][1]) )))
            elif est[0][0] == 'LinearSVC': transformer_list.append(('svc', LinearSVC(C=float(est[1][0][1]), class_weight=(est[1][1][1]) )))
            elif est[0][0] == 'KNeighborsClassifier': transformer_list.append(('knn_class', KNeighborsClassifier(n_neighbors=int(est[1][0][1]) )))
            elif est[0][0] == 'DecisionTreeClassifier': transformer_list.append(('tree_class', DecisionTreeClassifier(criterion=est[1][0][1], max_leaf_nodes=int(est[1][1][1]) )))

            # stworzenie Pipeline
            pipe = Pipeline(transformer_list)

            # odczytanie pliku wgranego przez użytkownika
            wb = pd.read_excel(io=request.session['filename'], header=0)

            # podział danych na x i y (kolumny cech i kolumnę etykiet)
            X = wb.iloc[:, : -1] 
            y = wb.iloc[:, -1]

            # wybór sposobu podziału danych na zbiór uczący i testujący

            # wybór podziału procentowego
            if ev[0] == 'Train test split':
                test_size = float(ev[1][0][1])
                X_train,X_test,y_train,y_test=train_test_split(X, y, test_size=test_size, random_state=0)              

                pipe.fit(X_train,y_train)
                y_pred=pipe.predict(X_test)
                pipe.score(X_test,y_test)

                if est[0][1] == 'reg':
                    
                    scores1_reg = {
                        'explained_variance_score': explained_variance_score(y_test, y_pred),
                        'r2_score': r2_score(y_test, y_pred),
                        'neg_mean_absolute_error': mean_absolute_error(y_test, y_pred),
                        'neg_mean_squared_error': mean_squared_error(y_test, y_pred) 
                    }

                    for x in scores1_reg.keys():
                        scores1_reg[x] = round(scores1_reg[x], 4)

                    context = {'scores1_reg': scores1_reg}

                elif est[0][1] == 'clf':

                    scores1_clf = {
                        'precision_macro': precision_score(y_test, y_pred, average='macro'),
                        'precision_micro': precision_score(y_test, y_pred, average='micro'),
                        'recall_macro': recall_score(y_test, y_pred, average='macro'),
                        'recall_micro': recall_score(y_test, y_pred, average='micro'),
                        'f1_macro': f1_score(y_test, y_pred, average='macro'),
                        'f1_micro': f1_score(y_test, y_pred, average='micro'),
                        'accuracy': accuracy_score(y_test, y_pred)
                    }

                    for x in scores1_clf.keys():
                        scores1_clf[x] = round(scores1_clf[x], 4)

                    context = {'scores1_clf': scores1_clf}

            # wybór podziału krzyżowo-walidacyjnego
            elif ev[0] == 'Cross validate':
                cv = int(ev[1][0][1])

                if est[0][1] == 'reg': 
                    scoring = ['explained_variance', 'r2', 'neg_mean_absolute_error', 'neg_mean_squared_error' ]

                    scores2_reg = cross_validate(pipe, X, y, cv=cv, scoring=scoring)

                    for x in scores2_reg.keys():
                        scores2_reg[x] = [scores2_reg[x].mean(), scores2_reg[x].std()]

                    for x, y in scores2_reg.items():
                        list_reg = []
                        for j in y:
                            list_reg.append(round(j, 4))
                        scores2_reg[x] = list_reg

                    context = {'scores2_reg': scores2_reg}

                elif est[0][1] == 'clf': 
                    scoring = ['precision_macro', 'precision_micro', 'recall_macro', 'recall_micro', 'f1_macro', 'f1_micro', 'accuracy']

                    scores2_clf = cross_validate(pipe, X, y, cv=cv, scoring=scoring)

                    for x in scores2_clf.keys():
                        scores2_clf[x] = [scores2_clf[x].mean(), scores2_clf[x].std()]

                    for x, y in scores2_clf.items():
                        list_clf = []
                        for j in y:
                            list_clf.append(round(j, 4))
                        scores2_clf[x] = list_clf

                    context = {'scores2_clf': scores2_clf}
            
            # możliwość zapisania modelu w bazie danych
            if request.method == 'POST':
                filename = request.POST['filename']
                file = "documents/user_" + str(request.user.id) + "/models/" + filename + ".pickle"
                existing_files = MLModel.objects.filter(file = file, user = request.user)
                if existing_files.count() == 0:
                    os.makedirs("documents/user_" + str(request.user.id) + "/models/", exist_ok=True)
                    with open(file, 'wb') as f:
                        pickle.dump(pipe, f)

                    newmodel = MLModel(file = file)
                    newmodel.user = request.user
                    newmodel.save()
                    messages.success(request, "Pomyślnie zapisano model!")
                    return redirect('models')
                elif existing_files.count() > 0:
                    messages.success(request, "Model o tej nazwie już istnieje")
            return render(request, 'compute.html', context)
        
        elif 'filename' not in request.session:
            messages.success(request, "Najpierw wybierz plik")
            return redirect('upload')  
        
        else:
            messages.success(request, "Dodaj estymator i metodę podziału danych")
            return render(request, 'compute.html', {}) 
        
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def show_models(request):
    if request.user.is_authenticated:
        models = MLModel.objects.filter(user=request.user)
        return render(request, 'models_list.html', {'models': models})
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def download_model(request, pk):
    if request.user.is_authenticated:
        to_download = get_object_or_404(MLModel, pk=pk)
        filename = str(to_download.file)
        if filename:
            with open(filename, 'rb') as fl:
                response = HttpResponse(fl.read(), content_type="application/octet-stream")
                response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(filename)
                messages.success(request, "Pomyślnie pobrano plik!")
                return response 
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def delete_model(request, pk):
    if request.user.is_authenticated:
        to_delete = get_object_or_404(MLModel, pk=pk)
        filename = str(to_delete.file)
        if filename:
            to_delete.delete()
            os.remove(filename)
            messages.success(request, "Pomyślnie usunięto plik!")
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def transformer_up(request):
    if request.user.is_authenticated:
        if 'my_pipeline' in request.session:
            transformer_list = request.session['my_pipeline']
            pos = int(request.GET.get('pos'))
            switch_up(transformer_list, pos)
            request.session['my_pipeline'] = transformer_list
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def transformer_down(request):
    if request.user.is_authenticated:
        if 'my_pipeline' in request.session:
            transformer_list = request.session['my_pipeline']
            pos = int(request.GET.get('pos'))
            switch_down(transformer_list, pos)
            request.session['my_pipeline'] = transformer_list
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def delete_step(request):
    if request.user.is_authenticated:
        if 'my_pipeline' in request.session:
            transformer_list = request.session['my_pipeline']
            if len(transformer_list) > 1:
                pos = int(request.GET.get('pos'))
                transformer_list.pop(pos)
                request.session['my_pipeline'] = transformer_list
            else:
                del request.session['my_pipeline']
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')
    
def delete_est(request):
    if request.user.is_authenticated:
        if 'my_estimator' in request.session:
            del request.session['my_estimator']
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')

def delete_ev(request):
    if request.user.is_authenticated:
        if 'my_evaluation' in request.session:
            del request.session['my_evaluation']
        return redirect('experiment')
    else:
        messages.success(request, "Zaloguj się")
        return redirect('home')